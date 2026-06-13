#!/usr/bin/env python3
"""
build_xlsx.py — Generate FORGE_INFO698_Gantt.xlsx from data/gantt.json.

Run from repo root:
    python scripts/build_xlsx.py

Reads the same source of truth as the website Gantt (data/gantt.json):
phase-grouped tasks with start/end week indices, milestones, and windows.
The resulting .xlsx is written to the repo root and committed alongside
gantt.json so external viewers (advisor, sponsor) get a familiar spreadsheet
with the bars rendered as filled week-columns.
"""

from __future__ import annotations

import json
import sys
from datetime import datetime, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Missing dependency: pip install openpyxl")


REPO_ROOT = Path(__file__).resolve().parent.parent
JSON_PATH = REPO_ROOT / "data" / "gantt.json"
OUT_PATH  = REPO_ROOT / "FORGE_INFO698_Gantt.xlsx"

# Phase fills (hex, no #) — match the site legend.
PHASE_FILLS = {
    "Intro":        "8893A8",
    "Planning":     "8893A8",
    "MVP P1":       "5E9B86",
    "MVP P2":       "5A7BA8",
    "Stretch":      "C67D3E",
    "Write-up":     "C79A3E",
    "Deliverables": "5BA56B",
}
PHASE_LABEL = {
    "Intro": "Intro / Planning", "Planning": "Intro / Planning",
    "MVP P1": "MVP — Phase 1 (standalone)",
    "MVP P2": "MVP — Phase 2 (integrate + deploy)",
    "Stretch": "MVP — Phase 3 / stretch",
    "Write-up": "Write-up / Deliverables", "Deliverables": "Turn in",
}
RISK_FILL = "F1DEC9"  # light tint for the risk-management window columns


def parse_week0(s: str):
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d")
    except ValueError:
        return None


def build():
    if not JSON_PATH.exists():
        sys.exit(f"gantt.json not found at {JSON_PATH}")

    data = json.loads(JSON_PATH.read_text())
    tasks = data.get("tasks", [])
    milestones = data.get("milestones", [])
    windows = data.get("windows", [])
    week0 = parse_week0(data.get("week0_start", ""))
    num_weeks = data.get("num_weeks") or (
        max((t.get("end", t.get("start", 0)) for t in tasks), default=0) + 1
    )

    # weeks covered by any risk window (for column tinting)
    risk_weeks = set()
    for w in windows:
        s = w.get("start", 0)
        e = w.get("end", s)
        risk_weeks.update(range(s, e + 1))

    wb = Workbook()
    ws = wb.active
    ws.title = "FORGE_Gantt"

    header_fill = PatternFill("solid", start_color="0F1F3D")
    white_bold  = Font(color="FFFFFF", bold=True, name="Calibri")
    normal      = Font(name="Calibri")
    bold        = Font(bold=True, name="Calibri")
    italic_gray = Font(italic=True, color="666666", name="Calibri")
    light_gray  = PatternFill("solid", start_color="F4F6FA")
    risk_tint   = PatternFill("solid", start_color=RISK_FILL)
    yellow_hl   = PatternFill("solid", start_color="FFF2CC")
    thin = Side(border_style="thin", color="D9DCE3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ---- parameter block ------------------------------------------------
    ws["A1"] = "Week 0 start:"; ws["A1"].font = bold
    ws["B1"] = data.get("week0_start", ""); ws["B1"].fill = yellow_hl
    ws["B1"].number_format = "yyyy-mm-dd"
    ws["C1"] = "(Edit data/gantt.json week0_start)"; ws["C1"].font = italic_gray

    ws["A2"] = "Project:"; ws["A2"].font = bold
    ws["B2"] = data.get("project", "FORGE"); ws["B2"].font = normal

    ws["A3"] = "Owner:"; ws["A3"].font = bold
    ws["B3"] = data.get("owner", ""); ws["B3"].font = normal

    ws["A4"] = "Updated:"; ws["A4"].font = bold
    ws["B4"] = data.get("last_updated", ""); ws["B4"].font = normal

    # ---- header ---------------------------------------------------------
    HEADER_ROW = 6
    headers = ["Phase", "Task", "Start", "End", "Status"]
    for w in range(num_weeks):
        headers.append(f"W{w}")
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=HEADER_ROW, column=col_idx, value=h)
        c.fill = header_fill
        c.font = white_bold
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    WEEK_COL0 = 6  # first week column (1-indexed) => F

    # ---- task rows ------------------------------------------------------
    start_row = HEADER_ROW + 1
    for i, t in enumerate(tasks):
        r = start_row + i
        phase = t.get("phase", "")
        s = t.get("start", 0)
        e = t.get("end", s)
        fill_hex = PHASE_FILLS.get(phase, "CCCCCC")
        phase_fill = PatternFill("solid", start_color=fill_hex)

        ws.cell(row=r, column=1, value=PHASE_LABEL.get(phase, phase)).font = normal
        ws.cell(row=r, column=2, value=t.get("name", "")).font = normal

        if week0:
            sd = week0 + timedelta(days=s * 7)
            ed = week0 + timedelta(days=e * 7 + 6)
            ws.cell(row=r, column=3, value=sd).number_format = "yyyy-mm-dd"
            ws.cell(row=r, column=4, value=ed).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=5, value=t.get("status", "planned")).alignment = Alignment(horizontal="center")

        for w in range(num_weeks):
            c = ws.cell(row=r, column=WEEK_COL0 + w)
            c.border = border
            if s <= w <= e:
                c.fill = phase_fill
            elif w in risk_weeks:
                c.fill = risk_tint
            else:
                c.fill = light_gray

        for col in range(1, 6):
            ws.cell(row=r, column=col).border = border

    last_task_row = start_row + len(tasks) - 1

    # ---- milestones row -------------------------------------------------
    ms_row = last_task_row + 1
    ws.cell(row=ms_row, column=2, value="Milestones (◈)").font = bold
    ws.cell(row=ms_row, column=1, value="").border = border
    ws.cell(row=ms_row, column=2).border = border
    for col in (3, 4, 5):
        ws.cell(row=ms_row, column=col).border = border
    ms_by_week = {}
    for j, m in enumerate(milestones, start=1):
        ms_by_week.setdefault(m["week"], []).append(str(j))
    for w in range(num_weeks):
        c = ws.cell(row=ms_row, column=WEEK_COL0 + w)
        c.border = border
        if w in ms_by_week:
            c.value = "◈" + "/".join(ms_by_week[w])
            c.font = Font(bold=True, color="0F1F3D", name="Calibri")
            c.alignment = Alignment(horizontal="center")
        elif w in risk_weeks:
            c.fill = risk_tint

    # ---- widths / freeze ------------------------------------------------
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    for w in range(num_weeks):
        ws.column_dimensions[get_column_letter(WEEK_COL0 + w)].width = 5
    ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=WEEK_COL0)

    # ---- phase legend ---------------------------------------------------
    legend_start = ms_row + 2
    ws.cell(row=legend_start, column=1, value="Phase legend").font = bold
    seen = set()
    li = 0
    for phase, hex_ in PHASE_FILLS.items():
        lab = PHASE_LABEL.get(phase, phase)
        if lab in seen:
            continue
        seen.add(lab)
        r = legend_start + 1 + li
        sw = ws.cell(row=r, column=1, value=""); sw.fill = PatternFill("solid", start_color=hex_); sw.border = border
        ws.cell(row=r, column=2, value=lab).font = normal
        li += 1
    # risk-window swatch
    r = legend_start + 1 + li
    sw = ws.cell(row=r, column=1, value=""); sw.fill = risk_tint; sw.border = border
    ws.cell(row=r, column=2, value="Risk-management window").font = normal

    # ---- milestone key --------------------------------------------------
    mk_start = legend_start + li + 3
    ws.cell(row=mk_start, column=1, value="Milestones").font = bold
    for j, m in enumerate(milestones, start=1):
        ws.cell(row=mk_start + j, column=1, value=f"◈{j}").font = bold
        ws.cell(row=mk_start + j, column=2, value=f"W{m['week']} — {m['label']}").font = normal

    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")


if __name__ == "__main__":
    build()
